"""
数据库服务模块
==============
提供 MySQL 数据库连接和 CRUD 操作。
按 docs/函数参数需求文档.md §2 实现所有函数签名。

依赖:
    - PyMySQL (pip install pymysql)
    - config.py (MYSQL_HOST, MYSQL_PORT, MYSQL_USER, MYSQL_PASSWORD, MYSQL_DATABASE)
    - models/user.py (User 类)

作者: 苏文韬
日期: 2026-06-08
"""

import pymysql
import json
from datetime import datetime, date
from typing import Optional, List, Dict, Any

from config import get_config
from models.user import User

config = get_config()


# ═══════════════════════════════════════════════════════════════
# §2.1 数据库连接
# ═══════════════════════════════════════════════════════════════

def get_db_connection() -> pymysql.Connection:
    """
    返回一个 MySQL 数据库连接。
    连接参数从 config.py 读取。

    返回:
        pymysql.Connection — 已建立连接的数据库对象

    注意:
        - 使用 autocommit=False，由调用方决定是否 commit
        - 连接使用完后调用方负责 close()
        - 使用 utf8mb4 字符集以支持中文和 Emoji
    """
    return pymysql.connect(
        host=config.MYSQL_HOST,
        port=config.MYSQL_PORT,
        user=config.MYSQL_USER,
        password=config.MYSQL_PASSWORD,
        database=config.MYSQL_DATABASE,
        charset="utf8mb4",
        cursorclass=pymysql.cursors.DictCursor,  # 返回字典格式，方便转 JSON
        autocommit=False,
    )


# ═══════════════════════════════════════════════════════════════
# §2.2 用户查询（认证用）
# ═══════════════════════════════════════════════════════════════

def get_user_by_username(username: str) -> Optional[User]:
    """
    根据用户名查询系统用户。

    参数:
        username: str — 用户名

    返回:
        User 对象（包含 user_id, username, password_hash, role, status）
        None — 用户不存在

    SQL:
        SELECT user_id, username, password, role, status, created_at
        FROM user WHERE username = %s
    """
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cursor:
            cursor.execute(
                "SELECT user_id, username, password, role, status, created_at "
                "FROM user WHERE username = %s",
                (username,),
            )
            row = cursor.fetchone()

        if row is None:
            return None

        return User(
            user_id=row["user_id"],
            username=row["username"],
            password_hash=row["password"],  # DB 列名 password → User 属性 password_hash
            role=row["role"],
            status=row["status"],
            created_at=row["created_at"],
        )
    except pymysql.MySQLError as e:
        raise RuntimeError(f"数据库查询失败 (get_user_by_username): {e}")
    finally:
        if conn:
            conn.close()


# ═══════════════════════════════════════════════════════════════
# §2.3 检查用户名是否存在
# ═══════════════════════════════════════════════════════════════

def check_user_exists(username: str) -> bool:
    """
    检查用户名是否已被注册。

    参数:
        username: str — 用户名

    返回:
        True — 已存在
        False — 不存在
    """
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cursor:
            cursor.execute(
                "SELECT 1 FROM user WHERE username = %s LIMIT 1",
                (username,),
            )
            return cursor.fetchone() is not None
    except pymysql.MySQLError as e:
        raise RuntimeError(f"数据库查询失败 (check_user_exists): {e}")
    finally:
        if conn:
            conn.close()


# ═══════════════════════════════════════════════════════════════
# §2.4 插入新用户
# ═══════════════════════════════════════════════════════════════

def insert_user(user: User) -> int:
    """
    向 user 表插入一条新用户记录。

    参数:
        user: User 对象（含 username, password_hash, role, status=1）

    返回:
        int — 新插入用户的 user_id（自增主键）

    SQL:
        INSERT INTO user (username, password, role, status, created_at)
        VALUES (%s, %s, %s, 1, NOW())
    """
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cursor:
            cursor.execute(
                "INSERT INTO user (username, password, role, status, created_at) "
                "VALUES (%s, %s, %s, %s, NOW())",
                (user.username, user.password_hash, user.role, user.status),
            )
        conn.commit()
        return cursor.lastrowid
    except pymysql.IntegrityError as e:
        if conn:
            conn.rollback()
        raise RuntimeError(f"用户插入失败（可能用户名已存在）: {e}")
    except pymysql.MySQLError as e:
        if conn:
            conn.rollback()
        raise RuntimeError(f"数据库插入失败 (insert_user): {e}")
    finally:
        if conn:
            conn.close()


# ═══════════════════════════════════════════════════════════════
# §2.5 多维度销售数据查询（UC-01）
# ═══════════════════════════════════════════════════════════════

def query_sales(
    start_date: str,
    end_date: str,
    region: Optional[str] = None,
    category_id: Optional[int] = None,
    channel: Optional[str] = None,
    group_by: str = "day",
) -> dict:
    """
    多维度销售数据查询（UC-01）。

    参数:
        start_date:  "2025-01-01"
        end_date:    "2025-12-31"
        region:      None=全国, "广东", ...
        category_id: None=全品类, 1, 2, ...
        channel:     None=全渠道, "PC", "Mobile", "Miniprogram"
        group_by:    "day" | "week" | "month"

    返回:
        {
            "summary": {
                "total_sales": 1234567.89,
                "total_orders": 5678,
                "avg_order_value": 217.45,
                "change_rate": 0.123
            },
            "series": [
                {"date": "2025-01-01", "sales": 12345.67, "orders": 56},
                ...
            ],
            "by_category": [
                {"category_name": "服装", "sales": 500000, "percentage": 0.40},
                ...
            ],
            "by_region": [
                {"region": "广东", "sales": 300000, "percentage": 0.24},
                ...
            ]
        }
    """
    # group_by 映射
    date_formats = {
        "day":   "DATE(order_date)",
        "week":  "DATE(DATE_SUB(order_date, INTERVAL WEEKDAY(order_date) DAY))",
        "month": "DATE_FORMAT(order_date, '%Y-%m-01')",
    }
    if group_by not in date_formats:
        raise ValueError(f"无效的 group_by 参数: {group_by}，可选值：day/week/month")

    date_expr = date_formats[group_by]

    conn = None
    try:
        conn = get_db_connection()

        # ── 构建动态 WHERE 条件 ──
        where_clauses = ["sr.order_date >= %s", "sr.order_date <= %s"]
        params: List[Any] = [start_date, f"{end_date} 23:59:59"]

        if region:
            where_clauses.append("sr.region = %s")
            params.append(region)
        if channel:
            where_clauses.append("sr.channel = %s")
            params.append(channel)

        join_clause = ""
        if category_id:
            # 如果指定了品类，需要 JOIN product 表
            join_clause = "JOIN product p ON sr.product_id = p.product_id"
            where_clauses.append("p.category_id = %s")
            params.append(category_id)

        where_sql = " AND ".join(where_clauses)

        # ── 1. 汇总统计 ──
        with conn.cursor() as cursor:
            summary_sql = f"""
                SELECT
                    COALESCE(SUM(sr.total_amount), 0) AS total_sales,
                    COUNT(*) AS total_orders,
                    COALESCE(ROUND(SUM(sr.total_amount) / NULLIF(COUNT(*), 0), 2), 0) AS avg_order_value
                FROM sales_record sr
                {join_clause}
                WHERE {where_sql}
            """
            cursor.execute(summary_sql, params)
            summary = cursor.fetchone()

        # ── 2. 时序序列 ──
        with conn.cursor() as cursor:
            series_sql = f"""
                SELECT
                    {date_expr} AS date,
                    COALESCE(SUM(sr.total_amount), 0) AS sales,
                    COUNT(*) AS orders
                FROM sales_record sr
                {join_clause}
                WHERE {where_sql}
                GROUP BY {date_expr}
                ORDER BY date
            """
            cursor.execute(series_sql, params)
            series = cursor.fetchall()
            # 日期序列化
            for item in series:
                if isinstance(item["date"], (date, datetime)):
                    item["date"] = item["date"].strftime("%Y-%m-%d")

        # ── 3. 按品类分布 ──
        by_category = []
        with conn.cursor() as cursor:
            cat_sql = f"""
                SELECT
                    COALESCE(c.category_name, '未分类') AS category_name,
                    COALESCE(SUM(sr.total_amount), 0) AS sales
                FROM sales_record sr
                JOIN product p ON sr.product_id = p.product_id
                JOIN category c ON p.category_id = c.category_id
                WHERE {where_sql}
                GROUP BY c.category_id, c.category_name
                ORDER BY sales DESC
            """
            cursor.execute(cat_sql, params)
            cat_rows = cursor.fetchall()
            total_sales = summary["total_sales"]
            for row in cat_rows:
                row["percentage"] = round(row["sales"] / total_sales, 4) if total_sales > 0 else 0.0
                row["sales"] = float(row["sales"])
            by_category = cat_rows

        # ── 4. 按地区分布 ──
        by_region = []
        with conn.cursor() as cursor:
            region_sql = f"""
                SELECT
                    sr.region,
                    COALESCE(SUM(sr.total_amount), 0) AS sales
                FROM sales_record sr
                {join_clause}
                WHERE {where_sql}
                GROUP BY sr.region
                ORDER BY sales DESC
            """
            cursor.execute(region_sql, params)
            reg_rows = cursor.fetchall()
            for row in reg_rows:
                row["percentage"] = round(row["sales"] / total_sales, 4) if total_sales > 0 else 0.0
                row["sales"] = float(row["sales"])
            by_region = reg_rows

        return {
            "summary": {
                "total_sales": float(summary["total_sales"]),
                "total_orders": int(summary["total_orders"]),
                "avg_order_value": float(summary["avg_order_value"]),
                "change_rate": 0.0,  # 环比变化率需要额外计算（对比上一周期），暂留 0
            },
            "series": series,
            "by_category": by_category,
            "by_region": by_region,
        }

    except pymysql.MySQLError as e:
        raise RuntimeError(f"销售数据查询失败 (query_sales): {e}")
    finally:
        if conn:
            conn.close()


# ═══════════════════════════════════════════════════════════════
# §2.6 报表导出（UC-09）
# ═══════════════════════════════════════════════════════════════

def export_report(
    report_type: str,
    params: dict,
    format: str,
) -> str:
    """
    生成报表并返回下载文件路径。

    参数:
        report_type: "daily" | "weekly" | "monthly" | "custom"
        params:      报表参数（时间范围、维度等），格式:
                     {
                         "start_date": "2025-01-01",
                         "end_date": "2025-12-31",
                         "region": null,
                         "category_id": null,
                         "channel": null,
                         "group_by": "day"
                     }
        format:      "excel" | "csv"（PDF 需要额外依赖，暂不支持）

    返回:
        str — 生成的文件路径（如 "/tmp/report_20260608_153000.xlsx"）
              调用方负责将此路径转为下载 URL

    注意:
        - Excel 导出需要 openpyxl 库: pip install openpyxl
        - 文件生成在系统临时目录
    """
    import os
    import tempfile

    # 参数校验
    valid_types = ("daily", "weekly", "monthly", "custom")
    if report_type not in valid_types:
        raise ValueError(f"无效的 report_type: {report_type}，可选值：{valid_types}")

    valid_formats = ("excel", "csv")
    if format not in valid_formats:
        raise ValueError(f"不支持的导出格式: {format}，可选值：{valid_formats}")

    # 先查询数据
    data = query_sales(
        start_date=params.get("start_date", "2020-01-01"),
        end_date=params.get("end_date", str(date.today())),
        region=params.get("region"),
        category_id=params.get("category_id"),
        channel=params.get("channel"),
        group_by=params.get("group_by", "day"),
    )

    # 生成文件名
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    ext = "xlsx" if format == "excel" else "csv"
    filename = f"report_{report_type}_{timestamp}.{ext}"
    filepath = os.path.join(tempfile.gettempdir(), filename)

    if format == "excel":
        _export_excel(data, filepath)
    elif format == "csv":
        _export_csv(data, filepath)

    return filepath


def _export_excel(data: dict, filepath: str) -> None:
    """将查询结果导出为 Excel 文件"""
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "Excel 导出需要 openpyxl 库。请执行: pip install openpyxl"
        )

    wb = openpyxl.Workbook()

    # ── Sheet 1: 汇总 ──
    ws_summary = wb.active
    ws_summary.title = "销售汇总"
    summary = data["summary"]
    ws_summary.append(["指标", "数值"])
    ws_summary.append(["总销售额（元）", summary["total_sales"]])
    ws_summary.append(["总订单量", summary["total_orders"]])
    ws_summary.append(["平均客单价（元）", summary["avg_order_value"]])

    # ── Sheet 2: 时序数据 ──
    ws_series = wb.create_sheet("时序趋势")
    if data["series"]:
        headers = list(data["series"][0].keys())
        ws_series.append(headers)
        for row in data["series"]:
            ws_series.append([row.get(h, "") for h in headers])

    # ── Sheet 3: 品类分布 ──
    ws_cat = wb.create_sheet("品类分布")
    if data["by_category"]:
        headers = list(data["by_category"][0].keys())
        ws_cat.append(headers)
        for row in data["by_category"]:
            ws_cat.append([row.get(h, "") for h in headers])

    # ── Sheet 4: 地区分布 ──
    ws_region = wb.create_sheet("地区分布")
    if data["by_region"]:
        headers = list(data["by_region"][0].keys())
        ws_region.append(headers)
        for row in data["by_region"]:
            ws_region.append([row.get(h, "") for h in headers])

    wb.save(filepath)


def _export_csv(data: dict, filepath: str) -> None:
    """将查询结果导出为 CSV 文件（仅导出时序数据）"""
    import csv

    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)

        # 写汇总
        writer.writerow(["=== 销售汇总 ==="])
        writer.writerow(["指标", "数值"])
        summary = data["summary"]
        for key, val in summary.items():
            writer.writerow([key, val])
        writer.writerow([])

        # 写时序
        writer.writerow(["=== 时序趋势 ==="])
        if data["series"]:
            headers = list(data["series"][0].keys())
            writer.writerow(headers)
            for row in data["series"]:
                writer.writerow([row.get(h, "") for h in headers])
